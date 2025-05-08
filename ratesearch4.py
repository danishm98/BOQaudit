import os
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from difflib import get_close_matches
import zipfile
import io

# Wide layout
st.set_page_config(layout="wide")
 
def search_phrase_in_excel(folder_path, search_phrase, exact_match):
    results = []
    for filename in os.listdir(folder_path):
        if (filename.endswith(".xlsx") or filename.endswith(".xls")) and not filename.startswith("~$"):
            file_path = os.path.join(folder_path, filename)
            try:
                if filename.endswith(".xlsx"):
                    df = pd.read_excel(file_path, engine='openpyxl', sheet_name=None)
                else:
                    df = pd.read_excel(file_path, engine='xlrd', sheet_name=None)

                for sheet_name, sheet_df in df.items():
                    sheet_df = sheet_df.loc[:, ~sheet_df.columns.str.contains('^Unnamed')]
                    sheet_df = sheet_df.dropna(axis=1, how='all')

                    if 'Item' in sheet_df.columns:
                        for index, row in sheet_df.iterrows():
                            cell_value = str(row['Item']).lower()
                            search_phrase_lower = search_phrase.lower()
                            if (exact_match and search_phrase_lower in cell_value) or \
                               (not exact_match and get_close_matches(search_phrase_lower, [cell_value])):
                                results.append({
                                    'No.': len(results) + 1,
                                    'File': filename,
                                    'file_path': file_path,
                                    'Search Phrase': search_phrase,
                                    'sheet_name': sheet_name,
                                    'row_index': index + 2,  # Excel row
                                    'row': row.iloc[:6].to_dict()
                                })
            except Exception as e:
                st.error(f"Error reading {file_path}: {e}")
    return results

def save_results_to_excel(results, search_phrase):
    output = io.BytesIO()
    rows = []

    for result in results:
        row_data = {
            'File': result['File'],
            'Search Phrase': result['Search Phrase'],
            'sheet_name': result['sheet_name'],
            'row_index': result['row_index']
        }
        row_data.update(result['row'])
        rows.append(row_data)

    df_results = pd.DataFrame(rows)
    df_results = df_results.loc[:, ~df_results.columns.str.contains('^Unnamed')]

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_results.to_excel(writer, index=False)
        workbook = writer.book
        sheet = writer.sheets['Sheet1']
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for column in sheet.iter_cols():
            if column[0].value and str(column[0].value).lower() == 'rate':
                for cell in column:
                    cell.fill = fill

        file_col_idx = df_results.columns.get_loc('File') + 1
        for row in range(2, len(df_results) + 2):
            file_cell = sheet[f'{get_column_letter(file_col_idx)}{row}']
            match_result = results[row - 2]
            file_name = match_result['File']
            sheet_name = match_result['sheet_name']
            row_number = match_result['row_index']
            static_path = r"C:\Users\MemonD\preprocessed BOQs minus NRM and Cat"
            full_path = os.path.join(static_path, file_name).replace("\\", "/")
            file_cell.hyperlink = f"file:///{full_path}#'{sheet_name}'!A{row_number}"
            file_cell.style = "Hyperlink"

        # Adjust column widths
        column_widths = {
            "File": 30,
            "Search Phrase": 20,
            "Item": 45,         # Reduced by 10%
            "Description": 45   # Reduced by 10%
        }

        for col_num, col_name in enumerate(df_results.columns, start=1):
            width = column_widths.get(col_name, 10)
            sheet.column_dimensions[get_column_letter(col_num)].width = width

    output.seek(0)
    return output

def display_results(results, search_phrase, folder_path):
    st.title(f"Search Results for '{search_phrase}'")
    st.write(f"Folder: {folder_path}")

    df_results = pd.DataFrame([{
        'File': result['File'],
        'Search Phrase': result['Search Phrase'],
        'sheet_name': result['sheet_name'],
        'row_index': result['row_index'],
        **result['row']
    } for result in results])

    st.dataframe(df_results, use_container_width=True)

    excel_data = save_results_to_excel(results, search_phrase)
    st.download_button(
        label="Download Results",
        data=excel_data,
        file_name=f"Search_Results_{search_phrase}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

def main():
    st.title("BOQ superset advanced search")

    uploaded_file = st.file_uploader("Upload a zipped folder of the BOQ Excel files", type=["zip"])
    search_phrase = st.text_input("Enter search phrase:")
    exact_match = st.radio("Match Type", ("Exact", "Approx")) == "Exact"

    if st.button("Search"):
        if uploaded_file and search_phrase:
            with zipfile.ZipFile(uploaded_file, 'r') as zip_ref:
                zip_ref.extractall("extracted_files")
            folder_path = "extracted_files"
            results = search_phrase_in_excel(folder_path, search_phrase, exact_match)
            if results:
                display_results(results, search_phrase, folder_path)
            else:
                st.warning("No results found.")
        else:
            st.error("Please upload a zipped folder and enter a search phrase.")

if __name__ == "__main__":
    main()
